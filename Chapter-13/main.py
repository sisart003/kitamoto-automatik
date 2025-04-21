import openpyxl, pprint
from openpyxl.utils import get_column_letter, column_index_from_string
# wb = openpyxl.load_workbook("Book1.xlsx")
# sheet = wb['Sheet3']    # The sheet we want to work with
# print(type(wb))
# print(wb.sheetnames) # The workbook's sheets' names
# print(sheet)
# print(type(sheet))
# print(sheet.title) # Get the sheet's title as a string
# anotherSheet = wb.active # Get the active sheet in the workbook
# print(anotherSheet)
# print(sheet['B1'].value)
# c = sheet['C2'] # Get the cell at B1
# print('Row %s, Column %s is %s' % (c.row, c.column, c.value)) # Get the row and column of a cell

# print(sheet.cell(row=1, column=2).value)
# for i in range(1, 8, 2):
    # print(i, sheet.cell(row=i, column=3).value)
# print(sheet.max_row)
# print(sheet.max_column)

# print(get_column_letter(900))
# print(get_column_letter(sheet.max_column)) # Get the column letter of the last column
# print(column_index_from_string('AA'))

# print(tuple(sheet['A1':'C4'])) # Get a tuple of tuples of the cells in the range A1:C4
# for rowOfCellObjects in sheet['A1':'C4']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- End of Row ---');

# print(list(sheet.columns)[1])
# for cellObj in list(sheet.columns)[2]:
#     print(cellObj.value) # Get the value of the cell in the second column

